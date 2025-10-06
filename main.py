import os
import sqlite3
import shutil
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
from datetime import datetime, timedelta
from cryptography import x509
from cryptography.hazmat.backends import default_backend
from cryptography.hazmat.primitives import serialization
from cryptography.hazmat.primitives.serialization import pkcs12
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

class CertificateOrganizerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Organizador de Certificados Digitais PFX")
        self.root.geometry("1000x700")
        
        # Configurações
        self.BASE_DIR = os.path.dirname(os.path.abspath(__file__))
        self.CERTIFICATES_DIR = os.path.join(self.BASE_DIR, 'certificates')
        self.VALID_DIR = os.path.join(self.CERTIFICATES_DIR, 'valid')
        self.EXPIRED_DIR = os.path.join(self.CERTIFICATES_DIR, 'expired')
        self.NEAR_EXPIRY_DIR = os.path.join(self.CERTIFICATES_DIR, 'near_expiry')
        self.INVALID_DIR = os.path.join(self.CERTIFICATES_DIR, 'invalid')
        
        self.DATABASE_PATH = os.path.join(self.BASE_DIR, 'certificates.db')
        self.SPREADSHEET_PATH = os.path.join(self.BASE_DIR, 'certificates.xlsx')
        self.NEAR_EXPIRY_DAYS = 30
        self.SUPPORTED_EXTENSIONS = {'.pfx', '.p12', '.pem', '.crt', '.cer', '.key'}
        self.EXCLUDE_DIRS = {'.git', '__pycache__', 'node_modules', 'venv', '.vscode'}
        
        # Variável para senha do PFX
        self.pfx_password = None
        
        self.create_directories()
        self.init_database()
        self.setup_ui()
        
    def create_directories(self):
        """Cria os diretórios necessários"""
        for directory in [self.VALID_DIR, self.EXPIRED_DIR, self.NEAR_EXPIRY_DIR, self.INVALID_DIR]:
            os.makedirs(directory, exist_ok=True)
    
    def init_database(self):
        """Inicializa o banco de dados com todas as colunas necessárias"""
        conn = sqlite3.connect(self.DATABASE_PATH)
        cursor = conn.cursor()
        
        # Verifica se a tabela existe e quais colunas tem
        cursor.execute("PRAGMA table_info(certificates)")
        existing_columns = [column[1] for column in cursor.fetchall()]
        
        # Colunas necessárias
        required_columns = [
            'id INTEGER PRIMARY KEY AUTOINCREMENT',
            'filename TEXT NOT NULL',
            'subject TEXT',
            'issuer TEXT',
            'valid_from DATE',
            'valid_to DATE',
            'serial_number TEXT',
            'status TEXT',
            'file_path TEXT',
            'original_path TEXT',
            'file_size INTEGER',
            'scan_date TIMESTAMP',
            'last_updated TIMESTAMP DEFAULT CURRENT_TIMESTAMP',
            'certificate_type TEXT',
            'has_private_key BOOLEAN',
            'notes TEXT'
        ]
        
        # Se a tabela não existe, cria com todas as colunas
        if not existing_columns:
            cursor.execute(f'''
                CREATE TABLE certificates (
                    {', '.join(required_columns)}
                )
            ''')
        else:
            # Adiciona colunas faltantes
            missing_columns = [
                'certificate_type', 'has_private_key'
            ]
            
            for column in missing_columns:
                if column not in existing_columns:
                    if column == 'has_private_key':
                        cursor.execute(f'ALTER TABLE certificates ADD COLUMN {column} BOOLEAN DEFAULT 0')
                    else:
                        cursor.execute(f'ALTER TABLE certificates ADD COLUMN {column} TEXT')
        
        conn.commit()
        conn.close()
    
    def setup_ui(self):
        """Configura a interface gráfica"""
        # Frame principal
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Configurar grid weights
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        main_frame.rowconfigure(5, weight=1)
        
        # Título
        title_label = ttk.Label(main_frame, text="Organizador de Certificados Digitais PFX", 
                               font=('Arial', 16, 'bold'))
        title_label.grid(row=0, column=0, columnspan=3, pady=(0, 20))
        
        # Seção de senha PFX
        password_frame = ttk.LabelFrame(main_frame, text="Configuração PFX/P12", padding="10")
        password_frame.grid(row=1, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10))
        password_frame.columnconfigure(1, weight=1)
        
        ttk.Label(password_frame, text="Senha PFX (opcional):").grid(row=0, column=0, sticky=tk.W, pady=2)
        
        self.password_var = tk.StringVar()
        password_entry = ttk.Entry(password_frame, textvariable=self.password_var, show="*", width=30)
        password_entry.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(5, 5), pady=2)
        
        ttk.Button(password_frame, text="Definir Senha", 
                  command=self.set_password).grid(row=0, column=2, padx=5, pady=2)
        
        ttk.Label(password_frame, text="Deixe em branco para tentar sem senha", 
                 font=('Arial', 8)).grid(row=1, column=1, sticky=tk.W, pady=2)
        
        # Seção de configurações
        config_frame = ttk.LabelFrame(main_frame, text="Configurações de Diretórios", padding="10")
        config_frame.grid(row=2, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10))
        config_frame.columnconfigure(1, weight=1)
        
        # Diretórios para escanear
        ttk.Label(config_frame, text="Diretórios para escanear:").grid(row=0, column=0, sticky=tk.W, pady=2)
        
        self.dirs_frame = ttk.Frame(config_frame)
        self.dirs_frame.grid(row=1, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=5)
        self.dirs_frame.columnconfigure(0, weight=1)
        
        self.dirs_listbox = tk.Listbox(self.dirs_frame, height=4)
        self.dirs_listbox.grid(row=0, column=0, sticky=(tk.W, tk.E))
        
        dirs_buttons_frame = ttk.Frame(self.dirs_frame)
        dirs_buttons_frame.grid(row=0, column=1, padx=(5, 0))
        
        ttk.Button(dirs_buttons_frame, text="Adicionar", 
                  command=self.add_directory).pack(fill=tk.X, pady=2)
        ttk.Button(dirs_buttons_frame, text="Remover", 
                  command=self.remove_directory).pack(fill=tk.X, pady=2)
        ttk.Button(dirs_buttons_frame, text="Limpar", 
                  command=self.clear_directories).pack(fill=tk.X, pady=2)
        
        # Diretórios padrão
        default_dirs = [
            os.path.expanduser('~'),
            os.path.join(self.BASE_DIR, 'certificates'),
            os.path.expanduser('~/Documents'),
            os.path.expanduser('~/Downloads')
        ]
        for directory in default_dirs:
            if os.path.exists(directory):
                self.dirs_listbox.insert(tk.END, directory)
        
        # Seção de operações
        ops_frame = ttk.LabelFrame(main_frame, text="Operações", padding="10")
        ops_frame.grid(row=3, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10))
        
        ttk.Button(ops_frame, text="Escanear e Organizar", 
                  command=self.scan_and_organize).grid(row=0, column=0, padx=5, pady=2)
        ttk.Button(ops_frame, text="Verificar Validade PFX", 
                  command=self.recursive_validity_check).grid(row=0, column=1, padx=5, pady=2)
        ttk.Button(ops_frame, text="Gerar Planilha", 
                  command=self.update_spreadsheet).grid(row=0, column=2, padx=5, pady=2)
        ttk.Button(ops_frame, text="Limpar Banco", 
                  command=self.clear_database).grid(row=0, column=3, padx=5, pady=2)
        
        # Barra de progresso
        self.progress = ttk.Progressbar(main_frame, mode='indeterminate')
        self.progress.grid(row=4, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=10)
        
        # Área de log
        log_frame = ttk.LabelFrame(main_frame, text="Log de Execução", padding="10")
        log_frame.grid(row=5, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 10))
        log_frame.columnconfigure(0, weight=1)
        log_frame.rowconfigure(0, weight=1)
        
        self.log_text = scrolledtext.ScrolledText(log_frame, height=15, width=100)
        self.log_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Status bar
        self.status_var = tk.StringVar(value="Pronto")
        status_bar = ttk.Label(main_frame, textvariable=self.status_var, relief=tk.SUNKEN)
        status_bar.grid(row=6, column=0, columnspan=3, sticky=(tk.W, tk.E))
    
    def set_password(self):
        """Define a senha para certificados PFX"""
        self.pfx_password = self.password_var.get().encode() if self.password_var.get() else None
        if self.pfx_password:
            self.log("Senha PFX definida")
        else:
            self.log("Senha PFX removida - tentando sem senha")
    
    def log(self, message):
        """Adiciona mensagem ao log"""
        self.log_text.insert(tk.END, f"{datetime.now().strftime('%H:%M:%S')} - {message}\n")
        self.log_text.see(tk.END)
        self.root.update_idletasks()
    
    def set_status(self, message):
        """Atualiza a barra de status"""
        self.status_var.set(message)
        self.root.update_idletasks()
    
    def add_directory(self):
        """Adiciona diretório à lista"""
        directory = filedialog.askdirectory()
        if directory and directory not in self.dirs_listbox.get(0, tk.END):
            self.dirs_listbox.insert(tk.END, directory)
    
    def remove_directory(self):
        """Remove diretório selecionado da lista"""
        selection = self.dirs_listbox.curselection()
        if selection:
            self.dirs_listbox.delete(selection[0])
    
    def clear_directories(self):
        """Limpa todos os diretórios da lista"""
        self.dirs_listbox.delete(0, tk.END)
    
    def get_selected_directories(self):
        """Retorna lista de diretórios selecionados"""
        return list(self.dirs_listbox.get(0, tk.END))
    
    def scan_directory(self, root_dir):
        """Escaneia recursivamente um diretório em busca de certificados"""
        certificate_files = []
        
        try:
            for root, dirs, files in os.walk(root_dir):
                # Remove diretórios excluídos
                dirs[:] = [d for d in dirs if d not in self.EXCLUDE_DIRS]
                
                for file in files:
                    if self._is_certificate_file(file):
                        full_path = os.path.join(root, file)
                        certificate_files.append(full_path)
                        
        except Exception as e:
            self.log(f"Erro ao escanear diretório {root_dir}: {str(e)}")
        
        return certificate_files
    
    def scan_multiple_directories(self, directories):
        """Escaneia múltiplos diretórios"""
        all_files = []
        for directory in directories:
            if os.path.exists(directory):
                self.log(f"Escaneando: {directory}")
                files = self.scan_directory(directory)
                all_files.extend(files)
                self.log(f"Encontrados {len(files)} certificados")
            else:
                self.log(f"Diretório não encontrado: {directory}")
        
        return all_files
    
    def _is_certificate_file(self, filename):
        """Verifica se o arquivo é um certificado suportado"""
        ext = os.path.splitext(filename)[1].lower()
        return ext in self.SUPPORTED_EXTENSIONS
    
    def get_file_info(self, file_path):
        """Obtém informações básicas do arquivo"""
        try:
            stat = os.stat(file_path)
            return {
                'file_path': file_path,
                'filename': os.path.basename(file_path),
                'directory': os.path.dirname(file_path),
                'size_bytes': stat.st_size,
                'modified_date': datetime.fromtimestamp(stat.st_mtime),
                'created_date': datetime.fromtimestamp(stat.st_ctime)
            }
        except Exception as e:
            self.log(f"Erro ao obter info do arquivo {file_path}: {str(e)}")
            return None
    
    def _analyze_pfx_certificate(self, file_path):
        """Analisa certificado no formato PFX/P12"""
        try:
            with open(file_path, 'rb') as f:
                pfx_data = f.read()
            
            # Tenta carregar o PFX com diferentes abordagens
            private_key = None
            certificate = None
            additional_certs = None
            
            # Primeiro tenta com senha se fornecida
            if self.pfx_password:
                try:
                    private_key, certificate, additional_certs = pkcs12.load_key_and_certificates(
                        pfx_data, 
                        self.pfx_password, 
                        backend=default_backend()
                    )
                except (ValueError, TypeError) as e:
                    self.log(f"  PFX com senha falhou: {str(e)}")
            
            # Se ainda não conseguiu, tenta sem senha
            if not certificate:
                try:
                    private_key, certificate, additional_certs = pkcs12.load_key_and_certificates(
                        pfx_data, 
                        None,  # Sem senha
                        backend=default_backend()
                    )
                except Exception as e:
                    return self._create_invalid_cert_info(file_path, f"PFX inválido: {str(e)}")
            
            if certificate:
                subject = certificate.subject.rfc4514_string()
                issuer = certificate.issuer.rfc4514_string()
                valid_from = certificate.not_valid_before
                valid_to = certificate.not_valid_after
                serial = str(certificate.serial_number)
                
                # Determina o status
                now = datetime.now()
                status = self._determine_status(valid_to, now)
                
                return {
                    'subject': subject,
                    'issuer': issuer,
                    'valid_from': valid_from.strftime('%Y-%m-%d'),
                    'valid_to': valid_to.strftime('%Y-%m-%d'),
                    'serial_number': serial,
                    'status': status,
                    'file_path': file_path,
                    'certificate_type': 'PFX/P12',
                    'has_private_key': private_key is not None,
                    'notes': 'PFX analisado com sucesso' + (' (com chave privada)' if private_key else ' (sem chave privada)')
                }
            else:
                return self._create_invalid_cert_info(file_path, "Nenhum certificado encontrado no arquivo PFX")
                
        except Exception as e:
            return self._create_invalid_cert_info(file_path, f"Erro na análise PFX: {str(e)}")
    
    def _analyze_standard_certificate(self, file_path):
        """Analisa certificados padrão (PEM, CRT, CER)"""
        try:
            with open(file_path, 'rb') as f:
                cert_data = f.read()
            
            cert = None
            has_private_key = False
            
            # Tenta como certificado PEM primeiro
            try:
                cert = x509.load_pem_x509_certificate(cert_data, default_backend())
            except:
                # Tenta como certificado DER
                try:
                    cert = x509.load_der_x509_certificate(cert_data, default_backend())
                except:
                    # Tenta como chave privada
                    try:
                        private_key = serialization.load_pem_private_key(
                            cert_data, 
                            password=None, 
                            backend=default_backend()
                        )
                        has_private_key = True
                        # Tenta extrair certificado da chave privada se possível
                        return self._create_invalid_cert_info(file_path, "Chave privada (sem certificado)")
                    except:
                        # Tenta como PKCS7 ou outros formatos
                        try:
                            # Tenta analisar como string para ver se é texto
                            try:
                                text_data = cert_data.decode('utf-8', errors='ignore')
                                if 'BEGIN CERTIFICATE' in text_data:
                                    # Tenta extrair o certificado PEM
                                    lines = text_data.split('\n')
                                    cert_start = None
                                    cert_end = None
                                    for i, line in enumerate(lines):
                                        if 'BEGIN CERTIFICATE' in line:
                                            cert_start = i
                                        if 'END CERTIFICATE' in line and cert_start is not None:
                                            cert_end = i
                                            break
                                    
                                    if cert_start is not None and cert_end is not None:
                                        pem_cert = '\n'.join(lines[cert_start:cert_end+1])
                                        cert = x509.load_pem_x509_certificate(pem_cert.encode(), default_backend())
                            except:
                                pass
                        except:
                            return self._create_invalid_cert_info(file_path, "Formato não reconhecido")
            
            if cert:
                subject = cert.subject.rfc4514_string()
                issuer = cert.issuer.rfc4514_string()
                valid_from = cert.not_valid_before
                valid_to = cert.not_valid_after
                serial = str(cert.serial_number)
                
                # Determina o status
                now = datetime.now()
                status = self._determine_status(valid_to, now)
                
                return {
                    'subject': subject,
                    'issuer': issuer,
                    'valid_from': valid_from.strftime('%Y-%m-%d'),
                    'valid_to': valid_to.strftime('%Y-%m-%d'),
                    'serial_number': serial,
                    'status': status,
                    'file_path': file_path,
                    'certificate_type': 'Standard',
                    'has_private_key': has_private_key,
                    'notes': 'Analisado com sucesso'
                }
            else:
                return self._create_invalid_cert_info(file_path, "Não foi possível analisar o certificado")
            
        except Exception as e:
            return self._create_invalid_cert_info(file_path, f"Erro na análise: {str(e)}")
    
    def _analyze_certificate(self, file_path):
        """Analisa o certificado baseado na extensão do arquivo"""
        ext = os.path.splitext(file_path)[1].lower()
        
        if ext in ['.pfx', '.p12']:
            return self._analyze_pfx_certificate(file_path)
        else:
            return self._analyze_standard_certificate(file_path)
    
    def _create_invalid_cert_info(self, file_path, reason):
        """Cria informações para certificados inválidos"""
        ext = os.path.splitext(file_path)[1].lower()
        cert_type = 'PFX/P12' if ext in ['.pfx', '.p12'] else 'Standard'
        
        return {
            'subject': 'INVALID',
            'issuer': 'INVALID', 
            'valid_from': '0000-00-00',
            'valid_to': '0000-00-00',
            'serial_number': 'INVALID',
            'status': 'invalid',
            'file_path': file_path,
            'certificate_type': cert_type,
            'has_private_key': False,
            'notes': reason
        }
    
    def _determine_status(self, valid_to, current_date):
        """Determina o status do certificado"""
        if isinstance(valid_to, str):
            return 'invalid'
        
        days_to_expire = (valid_to - current_date).days
        
        if days_to_expire < 0:
            return 'expired'
        elif days_to_expire <= self.NEAR_EXPIRY_DAYS:
            return 'near_expiry'
        else:
            return 'valid'
    
    def _get_destination_folder(self, status):
        """Retorna a pasta de destino baseado no status"""
        folders = {
            'valid': self.VALID_DIR,
            'expired': self.EXPIRED_DIR, 
            'near_expiry': self.NEAR_EXPIRY_DIR,
            'invalid': self.INVALID_DIR
        }
        return folders.get(status, self.INVALID_DIR)
    
    def _get_unique_filename(self, file_path):
        """Garante que o nome do arquivo seja único"""
        if not os.path.exists(file_path):
            return file_path
        
        base, ext = os.path.splitext(file_path)
        counter = 1
        while True:
            new_path = f"{base}_{counter}{ext}"
            if not os.path.exists(new_path):
                return new_path
            counter += 1
    
    def insert_certificate(self, cert_data):
        """Insere ou atualiza um certificado no banco de dados"""
        conn = sqlite3.connect(self.DATABASE_PATH)
        cursor = conn.cursor()
        
        cursor.execute('''
            INSERT OR REPLACE INTO certificates 
            (filename, subject, issuer, valid_from, valid_to, serial_number, status, 
             file_path, original_path, file_size, scan_date, certificate_type, has_private_key, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            cert_data['filename'],
            cert_data['subject'],
            cert_data['issuer'],
            cert_data['valid_from'],
            cert_data['valid_to'],
            cert_data['serial_number'],
            cert_data['status'],
            cert_data.get('file_path', ''),
            cert_data.get('original_path', ''),
            cert_data.get('file_size', 0),
            cert_data.get('scan_date', ''),
            cert_data.get('certificate_type', 'Unknown'),
            cert_data.get('has_private_key', False),
            cert_data.get('notes', '')
        ))
        
        conn.commit()
        conn.close()
    
    def scan_and_organize(self):
        """Escaneia diretórios e organiza certificados"""
        directories = self.get_selected_directories()
        if not directories:
            messagebox.showwarning("Aviso", "Selecione pelo menos um diretório para escanear")
            return
        
        self.progress.start()
        self.set_status("Escaneando e organizando certificados...")
        
        try:
            all_cert_files = self.scan_multiple_directories(directories)
            
            self.log(f"\nTotal de {len(all_cert_files)} certificados encontrados")
            
            organized_count = 0
            for file_path in all_cert_files:
                try:
                    file_info = self.get_file_info(file_path)
                    if not file_info:
                        continue
                    
                    cert_info = self._analyze_certificate(file_path)
                    destination = self._get_destination_folder(cert_info['status'])
                    
                    # Move o arquivo para a pasta apropriada
                    filename = file_info['filename']
                    new_path = os.path.join(destination, filename)
                    new_path = self._get_unique_filename(new_path)
                    
                    shutil.copy2(file_path, new_path)
                    
                    # Atualiza informações
                    cert_info['file_path'] = new_path
                    cert_info['filename'] = filename
                    cert_info['original_path'] = file_path
                    cert_info['file_size'] = file_info['size_bytes']
                    cert_info['scan_date'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    
                    # Salva no banco de dados
                    self.insert_certificate(cert_info)
                    
                    organized_count += 1
                    
                    # Log detalhado para PFX
                    if cert_info['certificate_type'] == 'PFX/P12':
                        key_info = "com chave privada" if cert_info['has_private_key'] else "sem chave privada"
                        self.log(f"✓ {filename} -> {cert_info['status']} (PFX {key_info}, Válido até: {cert_info['valid_to']})")
                    else:
                        self.log(f"✓ {filename} -> {cert_info['status']} (Válido até: {cert_info['valid_to']})")
                    
                except Exception as e:
                    self.log(f"✗ Erro ao processar {os.path.basename(file_path)}: {str(e)}")
            
            self.log(f"\nOrganização concluída: {organized_count}/{len(all_cert_files)} certificados processados")
            messagebox.showinfo("Sucesso", f"Organização concluída!\n{organized_count} certificados processados.")
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro durante a organização: {str(e)}")
        finally:
            self.progress.stop()
            self.set_status("Pronto")
    
    def recursive_validity_check(self):
        """Verifica validade de certificados em todos os diretórios especificados"""
        directories = self.get_selected_directories()
        if not directories:
            messagebox.showwarning("Aviso", "Selecione pelo menos um diretório para verificar")
            return
        
        self.progress.start()
        self.set_status("Verificando validade dos certificados PFX...")
        
        try:
            all_cert_files = self.scan_multiple_directories(directories)
            
            results = {
                'total': len(all_cert_files),
                'valid': 0,
                'expired': 0,
                'near_expiry': 0,
                'invalid': 0,
                'errors': 0,
                'pfx_with_key': 0
            }
            
            for file_path in all_cert_files:
                try:
                    cert_info = self._analyze_certificate(file_path)
                    results[cert_info['status']] += 1
                    
                    if cert_info.get('has_private_key'):
                        results['pfx_with_key'] += 1
                    
                    # Atualiza banco de dados
                    file_info = self.get_file_info(file_path)
                    if file_info:
                        cert_info['filename'] = file_info['filename']
                        cert_info['file_path'] = file_path
                        cert_info['file_size'] = file_info['size_bytes']
                        cert_info['scan_date'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                        
                        self.insert_certificate(cert_info)
                    
                    status_icon = {
                        'valid': '✓',
                        'near_expiry': '⚠',
                        'expired': '✗',
                        'invalid': '?'
                    }.get(cert_info['status'], '?')
                    
                    file_type = "PFX" if cert_info['certificate_type'] == 'PFX/P12' else "STD"
                    key_info = "🔑" if cert_info.get('has_private_key') else ""
                    
                    self.log(f"{status_icon} {file_type}{key_info} {os.path.basename(file_path):<25} {cert_info['status']:>12} Válido até: {cert_info['valid_to']}")
                    
                except Exception as e:
                    results['errors'] += 1
                    self.log(f"✗ Erro em {os.path.basename(file_path)}: {str(e)}")
            
            # Resumo
            self.log("\n" + "="*60)
            self.log("RESUMO DA VERIFICAÇÃO PFX")
            self.log("="*60)
            self.log(f"Total de certificados encontrados: {results['total']}")
            self.log(f"✓ Válidos: {results['valid']}")
            self.log(f"⚠ Próximos da expiração: {results['near_expiry']}")
            self.log(f"✗ Expirados: {results['expired']}")
            self.log(f"? Inválidos: {results['invalid']}")
            self.log(f"🔑 PFX com chave privada: {results['pfx_with_key']}")
            self.log(f"! Erros: {results['errors']}")
            self.log("="*60)
            
            messagebox.showinfo("Verificação Concluída", 
                              f"Verificação PFX concluída!\n\n"
                              f"Total: {results['total']} certificados\n"
                              f"Válidos: {results['valid']}\n"
                              f"Próximos da expiração: {results['near_expiry']}\n"
                              f"Expirados: {results['expired']}\n"
                              f"PFX com chave privada: {results['pfx_with_key']}")
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro durante a verificação: {str(e)}")
        finally:
            self.progress.stop()
            self.set_status("Pronto")
    
    def update_spreadsheet(self):
        """Atualiza a planilha Excel com os dados do banco"""
        self.progress.start()
        self.set_status("Gerando planilha...")
        
        try:
            # Obtém dados do banco
            conn = sqlite3.connect(self.DATABASE_PATH)
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM certificates ORDER BY valid_to DESC')
            certificates = cursor.fetchall()
            conn.close()
            
            # Cria/atualiza planilha
            try:
                wb = openpyxl.load_workbook(self.SPREADSHEET_PATH)
                ws = wb.active
                ws.delete_rows(2, ws.max_row)
            except:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Certificados PFX"
                
                # Cabeçalhos
                headers = ["ID", "Arquivo", "Subject", "Issuer", "Válido Desde", 
                          "Válido Até", "Número Série", "Status", "Tipo", 
                          "Chave Privada", "Caminho", "Caminho Original", 
                          "Tamanho", "Data Escaneamento", "Observações"]
                
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)
                    cell.alignment = Alignment(horizontal="center")
            
            # Preenche dados
            for row, cert in enumerate(certificates, 2):
                for col, value in enumerate(cert, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    
                    # Formatação condicional para status
                    if col == 8:  # Coluna de status
                        colors = {
                            'valid': '00FF00',
                            'near_expiry': 'FFFF00',
                            'expired': 'FF0000',
                            'invalid': '808080'
                        }
                        if value in colors:
                            cell.fill = PatternFill(
                                start_color=colors[value],
                                end_color=colors[value],
                                fill_type="solid"
                            )
            
            # Ajusta largura das colunas
            column_widths = {'A': 8, 'B': 20, 'C': 40, 'D': 40, 'E': 12, 
                           'F': 12, 'G': 20, 'H': 12, 'I': 10, 'J': 12,
                           'K': 30, 'L': 30, 'M': 10, 'N': 18, 'O': 30}
            
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            
            wb.save(self.SPREADSHEET_PATH)
            
            self.log(f"Planilha gerada: {self.SPREADSHEET_PATH}")
            messagebox.showinfo("Sucesso", f"Planilha gerada com sucesso!\n{self.SPREADSHEET_PATH}")
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao gerar planilha: {str(e)}")
        finally:
            self.progress.stop()
            self.set_status("Pronto")
    
    def clear_database(self):
        """Limpa o banco de dados"""
        if messagebox.askyesno("Confirmar", "Tem certeza que deseja limpar todo o banco de dados?"):
            try:
                conn = sqlite3.connect(self.DATABASE_PATH)
                cursor = conn.cursor()
                cursor.execute('DELETE FROM certificates')
                conn.commit()
                conn.close()
                
                self.log("Banco de dados limpo com sucesso")
                messagebox.showinfo("Sucesso", "Banco de dados limpo com sucesso!")
            except Exception as e:
                messagebox.showerror("Erro", f"Erro ao limpar banco de dados: {str(e)}")

def main():
    root = tk.Tk()
    app = CertificateOrganizerApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()